import os, docx, openpyxl; 
from openpyxl.styles import *;

direc = str(input("Directory: "))

store = {}

for f in os.listdir("./"+direc):
    doc = docx.Document("./"+direc+"/"+f)
    print(f)
    for line in doc.paragraphs:
        for run in line.runs:
            s = run.text.split()
            if len(s)>=3:
                if '(' in s[1] and ')' in s[2]:
                    b1 = s[1].index('(')
                    b2 = s[2].index(')')
                    char = s[0]
                    if "?" in char:
                        if(char[0] != '?'):
                            print(f"Unexpected Behaviour in {doc}: \"{char}\"")
                        char = char[1:]
                    if ":" in char:
                        char = char.replace(":", "")
                    pose = s[1][b1+1:]
                    expr = s[2][:b2]

                    if char not in store:
                      store[char] = {
                        pose: {
                          "total": 1,
                          "exprs": {
                            expr: 1
                          }
                        }
                      }
                      continue;

                    
                    if pose not in store[char]:
                      store[char][pose] = {
                        "total": 1,
                        "exprs": {
                          expr: 1
                        }
                      }
                      continue;

                    store[char][pose]["total"]+=1;

                    if expr not in store[char][pose]["exprs"]:
                      store[char][pose]["exprs"][expr] = 1
                      continue;

                    store[char][pose]["exprs"][expr]+=1

'''
OPENPYXL SECTION
'''

workbook = openpyxl.Workbook()

workbook.remove(workbook['Sheet'])

def next(char):
  return chr(ord(char)+1)

for char in store:
    poses = {k: v for k, v in sorted(store[char].items(), key=lambda o: -o[1]["total"])}
    sheet = workbook.create_sheet(title=char)
    sheet["A1"] = "POSES"
    sheet["A1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('96BAE6'))
    sheet["A2"] = "EXPRESSIONS"

    sheet.column_dimensions["A"].width = len("Expression")+4
    sheet.freeze_panes = sheet["B2"]
    cur = "B"

    for pose in poses:
        exprs = dict(sorted(store[char][pose]["exprs"].items(), key=lambda item: item[1],reverse=True))
        sheet[cur+"1"] = pose
        sheet[cur+"1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('C5D9F1'))
        sheet[next(cur)+"1"] = poses[pose]["total"]
        sheet[next(cur)+"1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('C5D9F1'))
        maxLen = len(pose)
        for expr, i in zip(exprs, range(1, 1+len(exprs))):
            sheet[cur+str(i+1)] = expr
            sheet[next(cur)+str(i+1)] = exprs[expr]
            maxLen = max(maxLen, len(expr))

        # set column width
        sheet.column_dimensions[cur].width = maxLen+2

        cur = next(next(cur))


workbook.save(filename=direc+" expressions.xlsx")
