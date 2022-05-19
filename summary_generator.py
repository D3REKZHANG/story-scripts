import openpyxl
from openpyxl.styles import PatternFill, Color

wb = openpyxl.load_workbook('character_expressions.xlsx')
ws = wb['Sheet1']

store = {}

for col in ws.iter_cols():
    char = col[0].value
    print(char)
    store[char] = {}
    for expr in col[1:]:
        if not expr.value or expr.value == "":
            break;
        spl = expr.value.split()
        if spl[0] in store[char]:
            store[char][spl[0]].append(spl[1])
        else:
            store[char][spl[0]] = [spl[1]]

workbook = openpyxl.Workbook()

workbook.remove(workbook['Sheet'])

for char in store:
    poses = {k: v for k, v in sorted(store[char].items(), key=lambda item: -len(item[1]))}
    sheet = workbook.create_sheet(title=char)
    sheet["A1"] = "POSES"
    sheet["A1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('96BAE6'))
    sheet["A2"] = "EXPRESSIONS"

    sheet.column_dimensions["A"].width = len("Expression")+4
    sheet.freeze_panes = sheet["B2"]
    cur = "B"

    for pose in poses:
        sheet[cur+"1"] = pose
        sheet[cur+"1"].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('C5D9F1'))
        maxLen = 0
        for expr, i in zip(poses[pose], range(2, 2+len(poses[pose]))):
            sheet[cur+str(i)] = expr
            maxLen = max(maxLen, len(expr))

        # set column width
        sheet.column_dimensions[cur].width = maxLen+2

        if cur[-1] == 'Z':
            cur = "A"*(len(cur)+1)
        else:
            cur = cur[1:]+chr(ord(cur[-1])+1)


workbook.save(filename="Summary.xlsx")
